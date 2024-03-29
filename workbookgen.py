from datetime import datetime
from openpyxl import Workbook
import os


def makeworkbook(projectname):
    if(not os.path.exists(os.path.join(os.getcwd(), "static", "generated"))):
        os.mkdir(os.path.join(os.getcwd(), "static", "generated"))
    current_wd = os.getcwd()
    pathtofile = os.path.join(current_wd, "static",
                              "generated", projectname+"WorkerAttendance.xlsx")
    if(os.path.exists(pathtofile)):
        os.remove(pathtofile)
    else:
        print(pathtofile, "file does not exist")
    wb = Workbook()
    ws = wb.active
    ws.title = projectname+"Worker Attendance"
    return wb


def find_in_jsondata(jsondata, name, date, type):
    if(type == "In Marked By"):
        for key, value in jsondata.items():
            if(value["date"] == date and value["Workername"] == name and "in" in value["type"].lower()):
                return value["Attendance-Marked-By"]
    if(type == "Out Marked By"):
        for key, value in jsondata.items():
            if(value["date"] == date and value["Workername"] == name and "out" in value["type"].lower()):
                return value["Attendance-Marked-By"]
    for key, value in jsondata.items():
        if(value["date"] == date and value["Workername"] == name and type.lower() in value["type"].lower()):
            return value["time"]
    return "Unavailable"


def movefile(filename):
    try:
        current_wd = os.getcwd()
        pathtofile = os.path.join(current_wd, filename)
        newpath = os.path.join(current_wd, "static", "generated", filename)
        os.rename(pathtofile, newpath)
    except Exception as e:
        print(e)


def add_data_to_workbook(data, wb, projectname, jsondata):
    # try:
    print("generating workbook")
    alldates = data[0]
    workernames = data[1]
    ws = wb.active
    row = 1
    col = 2
    ws.column_dimensions['A'].width = 30

    # marking dates
    print("generating dates")
    ws['A1'] = "Dates"
    for dates in alldates:
        ws.cell(row=row, column=col).value = dates
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+3)
        col += 4
    row += 1
    col = 2

    # marking days
    print("generating days")
    ws['A2'] = "Days"
    for dates in alldates:
        datetimeobj = datetime.strptime(dates, '%d-%m-%Y')
        day = datetimeobj.strftime('%A')
        ws.cell(row=row, column=col).value = day
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+3)
        col += 4
    row += 1
    col = 2

    # marking type
    print("generating type")
    ws['A3'] = "Type"
    for dates in alldates:
        ws.cell(row=row, column=col).value = "In"
        ws.cell(row=row, column=col+1).value = "In Marked By"
        ws.cell(row=row, column=col+2).value = "Out"
        ws.cell(row=row, column=col+3).value = "Out Marked By"
        col += 4

    # marking names
    print("generating names")
    ws['A4'] = "Names"
    row += 2
    for workers in workernames:
        ws.cell(row=row, column=1).value = workers
        row += 1

    row = 5
    col = 2
    for rownum in range(row, len(workernames)+row):
        name = ws.cell(row=rownum, column=1).value
        for colnum in range(col, (len(alldates)*4)+col):
            typeoftime = ws.cell(row=3, column=colnum).value
            date = ws.cell(row=1, column=colnum).value
            if(date == None):
                x = 1
                while(True):
                    date = ws.cell(row=1, column=colnum-x).value
                    if(date != None):
                        break
                    x += 1

            # print(name, date, typeoftime, rownum, colnum)
            print(typeoftime)
            to_write = find_in_jsondata(jsondata, name, date, typeoftime)
            ws.cell(row=rownum, column=colnum).value = to_write

        col = 2
    filename = projectname+"WorkerAttendance.xlsx"
    wb.save(filename)
    movefile(filename)
    print("generation finished")
    return True

    # except Exception as e:
    #     print(e)
    #     return False


def get_raw_data_for_workbook(data, projectname):
    alldates = list()
    workernames = list()
    for key, value in data.items():
        date = value["date"]
        workername = value["Workername"]
        if(date not in alldates):
            alldates.append(date)
        if(workername not in workernames):
            workernames.append(workername)

    # sort dates in ascending order dd-mm-yyyy
    alldates.sort(key=lambda x: datetime.strptime(x, '%d-%m-%Y'))
    workernames.sort()
    if(add_data_to_workbook([alldates, workernames], makeworkbook(projectname), projectname, data)):
        print("added data to workbook")
        return True
    else:
        print("failed to add data to workbook")
        return False


class employeeworkbook:

    def __init__(self) -> None:
        pass

    def get_all_employees(self, collection):
        self.employees = collection
        return self.make_workbook()

    def make_workbook(self):
        if(not os.path.exists(os.path.join(os.getcwd(), "static", "generated"))):
            os.mkdir(os.path.join(os.getcwd(), "static", "generated"))
        current_wd = os.getcwd()
        pathtofile = os.path.join(current_wd, "static",
                                  "generated", "EmployeeDetails.xlsx")
        if(os.path.exists(pathtofile)):
            os.remove(pathtofile)
        else:
            print(pathtofile, "file does not exist")
        wb = Workbook()
        ws = wb.active
        ws.title = "Employee Details"
        self.workbook = wb
        return self.header_scan()

    def header_scan(self):
        headers = []
        for x in self.employees:
            try:
                del self.employees[x]["_id"]
                del self.employees[x]["Password"]
            except:
                pass
            temp = list(self.employees[x].keys())
            for y in temp:
                if y not in headers:
                    headers.append(y)
        self.workbookheaders = headers
        return self.add_data_to_workbook()

    def add_data_to_workbook(self):
        try:
            ws = self.workbook.active
            headers = self.workbookheaders
            ws.append(headers)
            row = 1
            lenn = len(headers)
            for x in self.employees:
                row += 1
                for y in range(0, lenn):
                    try:
                        ws.cell(row=row, column=y +
                                1).value = self.employees[x][headers[y]]
                    except:
                        pass
            current_wd = os.getcwd()
            pathtofile = os.path.join(current_wd, "static",
                                      "generated", "EmployeeDetails.xlsx")
            self.workbook.save(pathtofile)
            return{"status": "success", "message": "file generated", "filepath": "../static/generated/EmployeeDetails.xlsx", "filename": "EmployeeDetails.xlsx"}
        except Exception as e:
            print(e)
            return{"status": "failure", "message": "file generation failed"}


class workerworkbook:

    def __init__(self, projectname) -> None:
        print("workerworkbook")
        self.projectname = projectname

    def get_all_workers(self, collection):
        self.workers = collection
        return self.make_workbook()

    def make_workbook(self):
        if(not os.path.exists(os.path.join(os.getcwd(), "static", "generated"))):
            os.mkdir(os.path.join(os.getcwd(), "static", "generated"))
        current_wd = os.getcwd()
        pathtofile = os.path.join(current_wd, "static",
                                  "generated", self.projectname+"WorkerDetails.xlsx")
        if(os.path.exists(pathtofile)):
            os.remove(pathtofile)
        else:
            print(pathtofile, "file does not exist")
        wb = Workbook()
        ws = wb.active
        ws.title = self.projectname+"Worker Details"
        self.workbook = wb
        return self.header_scan()

    def header_scan(self):
        headers = []
        for x in self.workers:
            try:
                del self.workers[x]["_id"]
                del self.workers[x]["Password"]
            except:
                pass
            temp = list(self.workers[x].keys())
            for y in temp:
                if y not in headers:
                    headers.append(y)
        print(headers)
        self.workbookheaders = headers
        return self.add_data_to_workbook()

    def add_data_to_workbook(self):
        try:
            ws = self.workbook.active
            headers = self.workbookheaders
            ws.append(headers)
            row = 1
            lenn = len(headers)
            for x in self.workers:
                row += 1
                for y in range(0, lenn):
                    try:
                        ws.cell(row=row, column=y +
                                1).value = self.workers[x][headers[y]]
                    except:
                        pass
            current_wd = os.getcwd()
            pathtofile = os.path.join(current_wd, "static","generated", self.projectname+"WorkerDetails.xlsx")
            self.workbook.save(pathtofile)
            return{"status": "success", "message": "file generated", "filepath": "../static/generated/"+self.projectname+"WorkerDetails.xlsx", "filename": self.projectname+"WorkerDetails.xlsx"}
        except Exception as e:
            print(e)
            return{"status": "failure", "message": "file generation failed"}
